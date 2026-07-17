from io import BytesIO, StringIO
import csv

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.dependencies import get_current_user, require_roles
from app.db.session import get_db
from app.models.teaching_class import AcademicTerm, ClassGroupMember, CourseSubject
from app.models.user import User
from app.schemas.common import ApiResponse
from app.schemas.teaching_class import (
    ClassGroupCreate, ClassGroupRead, ClassMaterialAdd, ClassStatusUpdate, ClassTeacherAdd,
    JoinClassRequest, JoinClassResult, JoinCodeUpdate, JoinRequestReview, RandomGroupCreate,
    RosterImportResult, SubjectCreate, TeachingClassCreate, TeachingClassRead, TermCreate,
)
from app.services.teaching_class_service import TeachingClassService


router = APIRouter(prefix="/teaching-classes", tags=["teaching-classes"])


@router.get("/subjects", response_model=ApiResponse[list[dict]])
def subjects(_: User = Depends(get_current_user), db: Session = Depends(get_db)) -> ApiResponse[list[dict]]:
    rows = db.scalars(select(CourseSubject).order_by(CourseSubject.name)).all()
    return ApiResponse(data=[{"id": row.id, "name": row.name, "code": row.code, "description": row.description} for row in rows])


@router.post("/subjects", response_model=ApiResponse[dict], status_code=status.HTTP_201_CREATED)
def create_subject(payload: SubjectCreate, _: User = Depends(require_roles("admin")),
                   db: Session = Depends(get_db)) -> ApiResponse[dict]:
    item = CourseSubject(name=payload.name.strip(), code=payload.code.strip().upper() if payload.code else None,
                         description=payload.description.strip() if payload.description else None)
    db.add(item); db.commit(); db.refresh(item)
    return ApiResponse(message="课程科目已创建", data={"id": item.id, "name": item.name, "code": item.code})


@router.get("/terms", response_model=ApiResponse[list[dict]])
def terms(_: User = Depends(get_current_user), db: Session = Depends(get_db)) -> ApiResponse[list[dict]]:
    rows = db.scalars(select(AcademicTerm).order_by(AcademicTerm.start_date.desc())).all()
    return ApiResponse(data=[{"id": row.id, "name": row.name, "start_date": row.start_date,
                              "end_date": row.end_date, "is_current": row.is_current} for row in rows])


@router.get("/teachers/available", response_model=ApiResponse[list[dict]])
def available_teachers(_: User = Depends(require_roles("teacher", "admin")),
                       db: Session = Depends(get_db)) -> ApiResponse[list[dict]]:
    rows = db.scalars(select(User).where(
        User.role == "teacher", User.approval_status == "approved"
    ).order_by(User.username)).all()
    return ApiResponse(data=[{"id": row.id, "username": row.username,
                              "identity_no": row.identity_no} for row in rows])


@router.post("/terms", response_model=ApiResponse[dict], status_code=status.HTTP_201_CREATED)
def create_term(payload: TermCreate, _: User = Depends(require_roles("admin")),
                db: Session = Depends(get_db)) -> ApiResponse[dict]:
    if payload.is_current:
        for current in db.scalars(select(AcademicTerm).where(AcademicTerm.is_current.is_(True))).all():
            current.is_current = False
    item = AcademicTerm(**payload.model_dump()); db.add(item); db.commit(); db.refresh(item)
    return ApiResponse(message="学期已创建", data={"id": item.id, "name": item.name})


@router.get("", response_model=ApiResponse[list[TeachingClassRead]])
def my_classes(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> ApiResponse[list[TeachingClassRead]]:
    return ApiResponse(data=TeachingClassService(db).list_for_user(user))


@router.post("", response_model=ApiResponse[TeachingClassRead], status_code=status.HTTP_201_CREATED)
def create_class(payload: TeachingClassCreate, user: User = Depends(require_roles("teacher", "admin")),
                 db: Session = Depends(get_db)) -> ApiResponse[TeachingClassRead]:
    service = TeachingClassService(db); item = service.create(user, payload)
    result = next(row for row in service.list_for_user(user) if row["id"] == item.id)
    return ApiResponse(message="教学班已创建", data=result)


@router.post("/join", response_model=ApiResponse[JoinClassResult])
def join_class(payload: JoinClassRequest, user: User = Depends(get_current_user),
               db: Session = Depends(get_db)) -> ApiResponse[JoinClassResult]:
    result = TeachingClassService(db).join(user, payload.join_code)
    return ApiResponse(message=result["message"], data=result)


@router.post("/{class_id}/leave", response_model=ApiResponse[dict])
def leave_class(class_id: int, user: User = Depends(get_current_user),
                db: Session = Depends(get_db)) -> ApiResponse[dict]:
    result = TeachingClassService(db).leave(user, class_id)
    return ApiResponse(message=result["message"], data=result)


@router.post("/join-requests/{request_id}/review", response_model=ApiResponse[dict])
def review_join_request(request_id: int, payload: JoinRequestReview,
                        user: User = Depends(require_roles("teacher", "admin")),
                        db: Session = Depends(get_db)) -> ApiResponse[dict]:
    item = TeachingClassService(db).review_request(user, request_id, payload.approved)
    return ApiResponse(message="申请已处理", data={"id": item.id, "status": item.status})


@router.get("/{class_id}/members", response_model=ApiResponse[list[dict]])
def class_members(class_id: int, user: User = Depends(require_roles("teacher", "admin")),
                  db: Session = Depends(get_db)) -> ApiResponse[list[dict]]:
    return ApiResponse(data=TeachingClassService(db).members(class_id, user))


@router.get("/{class_id}/join-requests", response_model=ApiResponse[list[dict]])
def class_requests(class_id: int, user: User = Depends(require_roles("teacher", "admin")),
                   db: Session = Depends(get_db)) -> ApiResponse[list[dict]]:
    return ApiResponse(data=TeachingClassService(db).pending_requests(class_id, user))


@router.put("/{class_id}/status", response_model=ApiResponse[dict])
def update_class_status(class_id: int, payload: ClassStatusUpdate,
                        user: User = Depends(require_roles("teacher", "admin")),
                        db: Session = Depends(get_db)) -> ApiResponse[dict]:
    item = TeachingClassService(db).update_status(class_id, user, payload.status)
    return ApiResponse(message="教学班状态已更新", data={"id": item.id, "status": item.status})


@router.put("/{class_id}/join-code", response_model=ApiResponse[dict])
def update_join_code(class_id: int, payload: JoinCodeUpdate,
                     user: User = Depends(require_roles("teacher", "admin")),
                     db: Session = Depends(get_db)) -> ApiResponse[dict]:
    item = TeachingClassService(db).update_join_code(class_id, user, payload.enabled, payload.regenerate)
    return ApiResponse(message="班级码已更新", data={"join_code": item.join_code,
                                                      "join_code_enabled": item.join_code_enabled})


@router.post("/{class_id}/teachers", response_model=ApiResponse[dict], status_code=status.HTTP_201_CREATED)
def add_class_teacher(class_id: int, payload: ClassTeacherAdd,
                      user: User = Depends(require_roles("teacher", "admin")),
                      db: Session = Depends(get_db)) -> ApiResponse[dict]:
    item = TeachingClassService(db).add_teacher(class_id, user, payload.user_id)
    return ApiResponse(message="协同教师已添加", data={"id": item.id, "user_id": item.user_id})


@router.post("/{class_id}/materials", response_model=ApiResponse[dict], status_code=status.HTTP_201_CREATED)
def add_class_material(class_id: int, payload: ClassMaterialAdd,
                       user: User = Depends(require_roles("teacher", "admin")),
                       db: Session = Depends(get_db)) -> ApiResponse[dict]:
    item = TeachingClassService(db).add_material(class_id, user, payload.course_id, payload.material_role)
    return ApiResponse(message="班级教材已更新", data={"id": item.id, "course_id": item.course_id,
                                                       "material_role": item.material_role})


def _read_roster(file: UploadFile, content: bytes) -> list[dict[str, str]]:
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        decoded = content.decode("utf-8-sig")
        rows = list(csv.DictReader(StringIO(decoded)))
    elif filename.endswith(".xlsx"):
        from openpyxl import load_workbook
        sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        rows = [dict(zip(headers, [str(value or "").strip() for value in row])) for row in values[1:]]
    else:
        raise HTTPException(status_code=400, detail="名单仅支持 CSV 或 XLSX")
    aliases = {"学号": "identity_no", "姓名": "student_name", "分组": "group_name",
               "identity_no": "identity_no", "student_name": "student_name", "group_name": "group_name"}
    return [{aliases.get(key, key): str(value or "").strip() for key, value in row.items()} for row in rows]


@router.post("/{class_id}/roster", response_model=ApiResponse[RosterImportResult])
async def import_roster(class_id: int, file: UploadFile = File(...),
                        user: User = Depends(require_roles("teacher", "admin")),
                        db: Session = Depends(get_db)) -> ApiResponse[RosterImportResult]:
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="名单文件不能超过 5 MB")
    rows = _read_roster(file, content)
    result = TeachingClassService(db).add_roster(class_id, user, rows)
    return ApiResponse(message="学生名单已导入", data=result)


@router.post("/{class_id}/groups", response_model=ApiResponse[ClassGroupRead], status_code=status.HTTP_201_CREATED)
def create_group(class_id: int, payload: ClassGroupCreate,
                 user: User = Depends(require_roles("teacher", "admin")),
                 db: Session = Depends(get_db)) -> ApiResponse[ClassGroupRead]:
    service = TeachingClassService(db); group = service.create_group(class_id, user, payload.name, payload.user_ids)
    ids = list(db.scalars(select(ClassGroupMember.user_id).where(ClassGroupMember.group_id == group.id)).all())
    return ApiResponse(message="学习小组已创建", data={"id": group.id, "name": group.name,
                                                       "sort_order": group.sort_order, "user_ids": ids})


@router.get("/{class_id}/groups", response_model=ApiResponse[list[ClassGroupRead]])
def list_groups(class_id: int, user: User = Depends(require_roles("teacher", "admin")),
                db: Session = Depends(get_db)) -> ApiResponse[list[ClassGroupRead]]:
    return ApiResponse(data=TeachingClassService(db).list_groups(class_id, user))


@router.post("/{class_id}/groups/random", response_model=ApiResponse[list[ClassGroupRead]])
def random_groups(class_id: int, payload: RandomGroupCreate,
                  user: User = Depends(require_roles("teacher", "admin")),
                  db: Session = Depends(get_db)) -> ApiResponse[list[ClassGroupRead]]:
    return ApiResponse(message="已完成随机分组",
                       data=TeachingClassService(db).random_groups(class_id, user, payload.group_count, payload.name_prefix))
