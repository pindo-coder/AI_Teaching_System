from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO, StringIO
import re
from pathlib import Path
from threading import BoundedSemaphore, Lock, Thread
from urllib.parse import urlparse

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import Engine, select
from sqlalchemy.orm import Session, sessionmaker

from app.models.material_import import MaterialImportBatch, MaterialImportItem
from app.models.user import User
from app.core.config import settings
from app.schemas.knowledge import (
    MaterialBatchCreate, MaterialBatchPreview, MaterialBatchRead, MaterialPreviewColumn,
    MaterialPreviewRow, MaterialPreviewSheet,
)
from app.services.material_center_service import MaterialCenterService


FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "source_url": ("网址", "链接", "url", "原文地址", "原文链接", "新闻链接", "详情页", "地址"),
    "source_title": ("标题", "文章标题", "新闻标题", "材料名称", "文章名称", "文件名", "名称"),
    "publisher": ("发布机关", "发布单位", "来源单位", "文章来源", "发布部门", "媒体", "来源", "机构"),
    "published_date": ("发布时间", "发布日期", "发文日期", "公开日期", "日期", "时间"),
    "applicable_scope": ("适用范围", "适用专题", "所属专题", "课程范围", "教学范围"),
    "version_label": ("版本标识", "版本号", "版本", "修订时间"),
    "knowledge_tags": ("知识点标签", "知识点", "主题词", "关键词", "标签"),
}


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\u3000", " ").split()).strip()


def _key(value: object) -> str:
    return re.sub(r"[\s_\-—:：/（）()]+", "", _text(value).lower())


def _looks_url(value: str) -> bool:
    parsed = urlparse(value.strip())
    return parsed.scheme == "https" and bool(parsed.netloc)


def _normal_date(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    match = re.search(r"(20\d{2})[年./\-](\d{1,2})[月./\-](\d{1,2})日?", value)
    if match:
        try:
            return date(*map(int, match.groups())).isoformat()
        except ValueError:
            return ""
    try:
        return date.fromisoformat(value[:10]).isoformat()
    except ValueError:
        return ""


def _decode_csv(content: bytes) -> list[list[object]]:
    decoded = None
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise HTTPException(status_code=400, detail="无法识别 CSV 文件编码")
    try:
        dialect = csv.Sniffer().sniff(decoded[:5000], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [list(row) for row in csv.reader(StringIO(decoded), dialect) if any(_text(item) for item in row)]


def _workbook_rows(filename: str, content: bytes) -> list[tuple[str, list[list[object]]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return [("CSV", _decode_csv(content))]
    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        return [(sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)]) for sheet in workbook.worksheets]
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="服务器尚未安装旧版 Excel 解析组件") from exc
        workbook = xlrd.open_workbook(file_contents=content)
        output: list[tuple[str, list[list[object]]]] = []
        for sheet in workbook.sheets():
            rows: list[list[object]] = []
            for row_index in range(sheet.nrows):
                row: list[object] = []
                for column_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, column_index)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        cell_value = xlrd.xldate_as_datetime(cell.value, workbook.datemode)
                    else:
                        cell_value = cell.value
                    row.append(cell_value)
                rows.append(row)
            output.append((sheet.name, rows))
        return output
    raise HTTPException(status_code=400, detail="仅支持 CSV、XLSX 和 XLS 文件")


def _header_index(rows: list[list[object]]) -> int:
    best_index, best_score = 0, -1
    for index, row in enumerate(rows[:20]):
        keys = [_key(value) for value in row]
        alias_hits = sum(any(alias in key for values in FIELD_ALIASES.values() for alias in values) for key in keys)
        next_rows = rows[index + 1:index + 6]
        url_hits = sum(_looks_url(_text(value)) for next_row in next_rows for value in next_row)
        score = alias_hits * 5 + min(url_hits, 5) + sum(bool(key) for key in keys) / 20
        if score > best_score:
            best_index, best_score = index, score
    return best_index


def _mapping(columns: list[str], data_rows: list[list[object]]) -> list[MaterialPreviewColumn]:
    output: list[MaterialPreviewColumn] = []
    used: set[int] = set()
    for field, aliases in FIELD_ALIASES.items():
        candidates: list[tuple[float, int]] = []
        for index, column in enumerate(columns):
            if index in used:
                continue
            key = _key(column)
            score = 0.0
            if key in aliases:
                score = 1.0
            elif any(alias in key or key in alias for alias in aliases if key):
                score = 0.88
            sample = [_text(row[index]) for row in data_rows[:30] if index < len(row) and _text(row[index])]
            if field == "source_url" and sample:
                score = max(score, sum(_looks_url(value) for value in sample) / len(sample) * 0.95)
            if field == "published_date" and sample:
                score = max(score, sum(bool(_normal_date(value)) for value in sample) / len(sample) * 0.78)
            candidates.append((score, index))
        score, index = max(candidates, default=(0.0, -1))
        if index >= 0 and score >= 0.48:
            used.add(index)
            output.append(MaterialPreviewColumn(field=field, column=columns[index], confidence=round(score, 2)))
    return output


def preview_material_file(filename: str, content: bytes) -> MaterialBatchPreview:
    sheets: list[MaterialPreviewSheet] = []
    for sheet_name, raw_rows in _workbook_rows(filename, content)[:10]:
        rows = [row for row in raw_rows if any(_text(item) for item in row)]
        if not rows:
            continue
        header_index = _header_index(rows)
        width = max(len(row) for row in rows)
        header = rows[header_index]
        columns: list[str] = []
        counts: dict[str, int] = {}
        for index in range(width):
            base = _text(header[index]) if index < len(header) else ""
            base = base or f"第{index + 1}列"
            counts[base] = counts.get(base, 0) + 1
            columns.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
        data_rows = rows[
            header_index + 1:header_index + 1 + settings.material_batch_max_items
        ]
        inferred = _mapping(columns, data_rows)
        field_columns = {item.field: columns.index(item.column) for item in inferred}
        preview_rows: list[MaterialPreviewRow] = []
        seen_urls: set[str] = set()
        for offset, row in enumerate(data_rows, start=header_index + 2):
            raw_data = {column: _text(row[index]) if index < len(row) else "" for index, column in enumerate(columns)}
            value = lambda field: _text(row[field_columns[field]]) if field in field_columns and field_columns[field] < len(row) else ""
            source_url = value("source_url").strip()
            published_raw = value("published_date")
            published_date = _normal_date(published_raw)
            errors: list[str] = []
            warnings: list[str] = []
            if not source_url:
                errors.append("缺少原文网址")
            elif not _looks_url(source_url):
                errors.append("网址必须是 HTTPS 公网地址")
            elif len(source_url) > 1000:
                errors.append("原文网址超过1000字符，请使用规范原文地址")
            elif source_url in seen_urls:
                warnings.append("文件内重复网址")
            if not value("source_title"):
                warnings.append("标题将尝试从网页识别")
            if not value("publisher"):
                warnings.append("发布机关将尝试从网页识别")
            if not published_date:
                warnings.append("日期将尝试从网页识别" if not published_raw else "日期格式需核对")
            seen_urls.add(source_url)
            tags = [item.strip() for item in re.split(r"[，,、;；|\n]", value("knowledge_tags")) if item.strip()]
            preview_rows.append(MaterialPreviewRow(
                row_number=offset, selected=not errors and not warnings, source_url=source_url,
                source_title=value("source_title")[:255], publisher=value("publisher")[:255],
                published_date=published_date, applicable_scope=value("applicable_scope")[:500],
                version_label=value("version_label")[:100], knowledge_tags=tags[:30],
                raw_data=raw_data, errors=errors, warnings=warnings,
            ))
        sheets.append(MaterialPreviewSheet(
            name=sheet_name, header_row=header_index + 1, columns=columns,
            mapping=inferred, rows=preview_rows,
        ))
    if not sheets:
        raise HTTPException(status_code=400, detail="文件中没有可预览的数据")
    return MaterialBatchPreview(filename=filename[:255], sheets=sheets)


class MaterialBatchService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def create(self, user: User, payload: MaterialBatchCreate) -> MaterialImportBatch:
        urls = [item.source_url.strip() for item in payload.items]
        if len(urls) != len(set(urls)):
            raise HTTPException(status_code=400, detail="提交数据中包含重复网址")
        if any(not _looks_url(url) for url in urls):
            raise HTTPException(status_code=400, detail="批量数据中包含非 HTTPS 网址")
        MaterialCenterService(self.db)._validate_scopes(
            user, "central", payload.course_ids, payload.chapter_ids, []
        )
        batch = MaterialImportBatch(
            created_by=user.id, original_filename=payload.original_filename,
            sheet_name=payload.sheet_name, status="queued", total_count=len(payload.items),
            course_ids=list(dict.fromkeys(payload.course_ids)),
            chapter_ids=list(dict.fromkeys(payload.chapter_ids)), access_policy=payload.access_policy,
        )
        self.db.add(batch); self.db.flush()
        self.db.add_all([
            MaterialImportItem(
                batch_id=batch.id, row_number=item.row_number, source_url=item.source_url.strip(),
                source_title=item.source_title.strip() if item.source_title else None,
                publisher=item.publisher.strip() if item.publisher else None,
                published_date=item.published_date, applicable_scope=item.applicable_scope,
                version_label=item.version_label, knowledge_tags=item.knowledge_tags,
                raw_data=item.raw_data, status="queued",
            ) for item in payload.items
        ])
        self.db.commit(); self.db.refresh(batch)
        return batch

    def require(self, batch_id: int, user: User) -> MaterialImportBatch:
        batch = self.db.get(MaterialImportBatch, batch_id)
        if batch is None:
            raise HTTPException(status_code=404, detail="批量导入任务不存在")
        if user.role != "admin" or batch.created_by != user.id:
            raise HTTPException(status_code=403, detail="无权查看该批量导入任务")
        return batch

    def read(self, batch: MaterialImportBatch) -> MaterialBatchRead:
        items = list(self.db.scalars(select(MaterialImportItem).where(
            MaterialImportItem.batch_id == batch.id
        ).order_by(MaterialImportItem.row_number)).all())
        return MaterialBatchRead.model_validate(batch).model_copy(update={"items": items})

    def list(self, user: User, limit: int = 30) -> list[MaterialImportBatch]:
        if user.role != "admin":
            raise HTTPException(status_code=403, detail="只有管理员可以查看批量导入任务")
        return list(self.db.scalars(select(MaterialImportBatch).where(
            MaterialImportBatch.created_by == user.id
        ).order_by(
            MaterialImportBatch.created_time.desc(), MaterialImportBatch.id.desc()
        ).limit(limit)).all())


_batch_slots = BoundedSemaphore(max(1, settings.material_batch_worker_concurrency))
_active_batch_ids: set[int] = set()
_active_batch_lock = Lock()


def process_material_batch(batch_id: int, bind: Engine) -> None:
    """执行一个持久化批次；并发数受配置限制，页面关闭不会影响该线程。"""

    with _active_batch_lock:
        if batch_id in _active_batch_ids:
            return
        _active_batch_ids.add(batch_id)
    try:
        with _batch_slots:
            _process_material_batch(batch_id, bind)
    finally:
        with _active_batch_lock:
            _active_batch_ids.discard(batch_id)


def _process_material_batch(batch_id: int, bind: Engine) -> None:
    factory = sessionmaker(bind=bind, autoflush=False, expire_on_commit=False)
    with factory() as db:
        batch = db.get(MaterialImportBatch, batch_id)
        if batch is None:
            return
        user = db.get(User, batch.created_by)
        if user is None:
            batch.status = "failed"; db.commit(); return
        batch.status = "processing"; db.commit()
        items = list(db.scalars(select(MaterialImportItem).where(
            MaterialImportItem.batch_id == batch.id,
            MaterialImportItem.status.in_(["queued", "failed"]),
        ).order_by(MaterialImportItem.row_number)).all())
        for item in items:
            item.status = "processing"; item.error_message = None; db.commit()
            try:
                document = MaterialCenterService(db).ingest_url(
                    user, source_url=item.source_url, source_title=item.source_title,
                    publisher=item.publisher, published_date=item.published_date,
                    applicable_scope=item.applicable_scope, version_label=item.version_label,
                    supersedes_document_id=None, access_policy=batch.access_policy,
                    course_ids=batch.course_ids, chapter_ids=batch.chapter_ids,
                    knowledge_tags=item.knowledge_tags,
                )
                item.document_id = document.id
                item.source_title = document.source_title
                item.publisher = document.publisher
                item.published_date = document.published_date
                item.status = "pending_review"
            except HTTPException as exc:
                item.status = "duplicate" if exc.status_code == 409 else "failed"
                item.error_message = str(exc.detail)[:1000]
                db.rollback()
                item = db.get(MaterialImportItem, item.id)
                if item is not None:
                    item.status = "duplicate" if exc.status_code == 409 else "failed"
                    item.error_message = str(exc.detail)[:1000]
            except Exception as exc:  # pragma: no cover - external website/network failures vary
                db.rollback()
                item = db.get(MaterialImportItem, item.id)
                if item is not None:
                    item.status = "failed"
                    item.error_message = f"处理失败：{str(exc)[:900]}"
            db.commit()
            completed = list(db.scalars(select(MaterialImportItem).where(
                MaterialImportItem.batch_id == batch.id
            )).all())
            batch = db.get(MaterialImportBatch, batch.id)
            batch.completed_count = sum(value.status not in {"queued", "processing"} for value in completed)
            batch.success_count = sum(value.status == "pending_review" for value in completed)
            batch.failed_count = sum(value.status == "failed" for value in completed)
            batch.duplicate_count = sum(value.status == "duplicate" for value in completed)
            db.commit()
        batch = db.get(MaterialImportBatch, batch_id)
        batch.status = "completed"
        db.commit()


def schedule_material_batch(batch_id: int, bind: Engine) -> None:
    """在守护线程中执行任务，使 HTTP 响应可以立即返回。"""

    Thread(
        target=process_material_batch,
        args=(batch_id, bind),
        name=f"material-batch-{batch_id}",
        daemon=True,
    ).start()


def recover_material_batches(bind: Engine) -> int:
    """服务启动时恢复排队中或被进程重启打断的批次。"""

    factory = sessionmaker(bind=bind, autoflush=False, expire_on_commit=False)
    with factory() as db:
        batches = list(db.scalars(select(MaterialImportBatch).where(
            MaterialImportBatch.status.in_(["queued", "processing"])
        ).order_by(MaterialImportBatch.id)).all())
        batch_ids = [batch.id for batch in batches]
        if batch_ids:
            processing_items = db.scalars(select(MaterialImportItem).where(
                MaterialImportItem.batch_id.in_(batch_ids),
                MaterialImportItem.status == "processing",
            )).all()
            for item in processing_items:
                item.status = "queued"
                item.error_message = None
            for batch in batches:
                batch.status = "queued"
            db.commit()
    for batch_id in batch_ids:
        schedule_material_batch(batch_id, bind)
    return len(batch_ids)
